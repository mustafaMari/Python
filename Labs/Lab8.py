import argparse
import csv
import sys
import openpyxl
import openpyxl.styles


class row_handler:
    def __init__(self, row):
        self.city = row['City']
        self.state = row['State']
        self.victimAge = int(row['Victim Age'])
        self.year = row['Year']

    def __str__(self):
        print(f"The state: {self.state},\n"
              f"The city: {self.city},\n"
              f"The victim age: {self.victimAge},\n"
              f"The year of the incident: {self.year}")


def read_csv(file_name):
    list_content = []
    try:
        with open(file_name) as f:
            dictReader = csv.DictReader(f)
            for row in dictReader:
                list_content.append(row_handler(row))

        return list_content
    except FileNotFoundError:
        print("the dataset does not exist")
        sys.exit(0)


def number_crime_by_state():
    Dict = {}
    for data in dataset:
        if data.state not in Dict:
            counter = 1
            Dict[data.state] = {
                "number_of_victims": counter
            }
        else:
            Dict[data.state]["number_of_victims"] += 1
    return Dict


def number_of_cities():
    city_Set = set()
    for data in dataset:
        city_Set.add(data.city)
    return len(city_Set)


def average_age_of_victims():
    totalAge = 0
    for data in dataset:
        totalAge += data.victimAge
    return totalAge / len(dataset)


def save_to_excel(file_name):
    f_title = openpyxl.styles.Font(color="FF0000", italic=True, bold=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Homicide Crimes Statistics"
    average_age = sheet.cell(row=1, column=3)
    average_age.value = "Average age of victims "
    average_age.font = f_title
    average_age_value = sheet.cell(row=2, column=3)
    average_age_value.value = average_age_of_victims()
    cities_number = sheet.cell(row=1, column=4)
    cities_number.value = "Total Number of Cities"
    cities_number.font = f_title
    cities_number_value = sheet.cell(row=2, column=4)
    cities_number_value.value = number_of_cities()
    sheet.merge_cells('A1:B1')
    number_by_state = sheet.cell(row=1, column=1)
    number_by_state.value = "Number of Victims By State"
    number_by_state.font = f_title
    number_by_state = sheet.cell(row=2, column=1)
    number_by_state.value = "State"
    number_by_state = sheet.cell(row=2, column=2)
    number_by_state.value = "Number"
    counter = 3
    for data in number_crime_by_state():
        sheet.cell(row=counter, column=1, value=data)
        sheet.cell(row=counter, column=2, value=number_crime_by_state()[data]['number_of_victims'])
        counter += 1
    workbook.save(file_name)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="The idea of the application is to analyse data from CSV File,"
                                                 " perform few operations and have the option"
                                                 " to store the result of such operations to an Exel File")
    parser.add_argument('csv_file', type=str, help="The name of the CSV File")
    parser.add_argument('-o', help="To save results of all operations in the Exel File")
    args = parser.parse_args()
    name_csv_file = args.csv_file
    if name_csv_file[-4:] != '.csv':
        print("The file has a wrong format")
        sys.exit(0)
    else:
        dataset = read_csv(name_csv_file)
    if 'o' in args and args.o is not None:
        name_xlsx_file = args.o
        save_to_excel(name_xlsx_file)
    else:
        print(f"Number of cities in USA where there are reported victims is: {number_of_cities()}")
